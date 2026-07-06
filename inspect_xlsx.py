import os
import openpyxl

file_path = "inStuffGames 2020 (4).xlsx"

if not os.path.exists(file_path):
    print("Excel file not found!")
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheet Names:", wb.sheetnames)

# We want the sheet that has the games list. Usually it's "Inventory" or "Sales".
# Let's check sheet names.
sheet_name = None
for name in wb.sheetnames:
    if "inventory" in name.lower() or "sales" in name.lower() or "key" in name.lower():
        sheet_name = name
        break

if not sheet_name:
    # Default to the second sheet if first is Dashboard
    if len(wb.sheetnames) > 1:
        sheet_name = wb.sheetnames[1]
    else:
        sheet_name = wb.sheetnames[0]

print(f"Inspecting Sheet: {sheet_name}")
sheet = wb[sheet_name]
print(f"Dimensions: {sheet.dimensions}")

# Print headers from Row 1
headers = [cell.value for cell in sheet[1]]
print("Row 1 Headers:", headers)
# Sometimes headers are on Row 2 or 3
headers2 = [cell.value for cell in sheet[2]]
print("Row 2 Headers:", headers2)

# We want to search for row 8976 in the sheet (in openpyxl, sheet.cell is 1-indexed)
# Let's print rows 8970 to 8985
print("\n--- Inspecting Rows 8970 to 8982 ---")
for r_idx in range(8970, min(8985, sheet.max_row + 1)):
    title_cell = sheet.cell(row=r_idx, column=2)
    title = title_cell.value
    row_vals = [sheet.cell(row=r_idx, column=col).value for col in range(1, 15)]
    closed_date_cell = sheet.cell(row=r_idx, column=5)
    print(f"Row {r_idx}: {row_vals}")
    print(f"  Closed Date cell value: {closed_date_cell.value}, type: {type(closed_date_cell.value)}, number_format: {closed_date_cell.number_format}")
